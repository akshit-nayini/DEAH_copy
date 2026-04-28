"""
storage/local_storage.py
------------------------
Helpers for saving and loading test case / validation result files
to the local filesystem (CSV + Excel).
"""

from __future__ import annotations
import re
from datetime import datetime
from pathlib import Path

import pandas as pd


# ── Stem builders ─────────────────────────────────────────────────────────────

def build_stem(icd_filename: str | None, agent: str) -> str:
    """
    Build the output filename stem from the ICD filename and agent label.

    Examples
    --------
    icd_filename = 'model_SCRUM-42_20260411_21_mapping.csv', agent = 'testcases'
    → 'SCRUM-42_20260411_21_testcases_20260414_16'

    icd_filename = None, agent = 'testcases'
    → 'testcases_20260414_16'
    """
    now_dthr = datetime.now().strftime("%Y%m%d_%H")
    if icd_filename:
        m = re.search(r'(SCRUM-\d+)_(\d{8}_\d{2})', str(icd_filename), re.IGNORECASE)
        if m:
            return f"{m.group(1).upper()}_{m.group(2)}_{agent}_{now_dthr}"
    return f"{agent}_{now_dthr}"


def build_validator_stem(testcases_filename: str | None) -> str:
    """
    Build the validator stem from the generator output filename.

    Examples
    --------
    testcases_filename = 'SCRUM-42_20260411_21_testcases_20260414_16.csv'
    → 'SCRUM-42_20260411_21_validator_20260414_17'

    testcases_filename = None
    → 'validator_20260414_17'
    """
    now_dthr = datetime.now().strftime("%Y%m%d_%H")
    if testcases_filename:
        m = re.search(r'(SCRUM-\d+_\d{8}_\d{2})_testcases', str(testcases_filename), re.IGNORECASE)
        if m:
            return f"{m.group(1).upper()}_validator_{now_dthr}"
    return f"validator_{now_dthr}"


# ── Save helpers ──────────────────────────────────────────────────────────────

def save_csv(records: list[dict], output_dir: Path,
             prefix: str = "output", stem: str | None = None) -> Path:
    output_dir.mkdir(parents=True, exist_ok=True)
    if stem:
        path = output_dir / f"{stem}.csv"
    else:
        ts   = datetime.now().strftime("%Y%m%d_%H%M%S")
        path = output_dir / f"{prefix}_{ts}.csv"
    pd.DataFrame(records).to_csv(path, index=False)
    return path


def save_excel(records: list[dict], output_dir: Path,
               prefix: str = "output", stem: str | None = None,
               sheet_name: str = "Results", verdict_col: str | None = "verdict") -> Path:
    from openpyxl.styles import Font, PatternFill

    output_dir.mkdir(parents=True, exist_ok=True)
    if stem:
        path = output_dir / f"{stem}.xlsx"
    else:
        ts   = datetime.now().strftime("%Y%m%d_%H%M%S")
        path = output_dir / f"{prefix}_{ts}.xlsx"

    df = pd.DataFrame(records)
    with pd.ExcelWriter(path, engine="openpyxl") as writer:
        df.to_excel(writer, index=False, sheet_name=sheet_name)
        ws = writer.sheets[sheet_name]

        # Conditional formatting on verdict column
        if verdict_col and verdict_col in df.columns:
            green  = PatternFill("solid", fgColor="C6EFCE")
            red    = PatternFill("solid", fgColor="FFC7CE")
            yellow = PatternFill("solid", fgColor="FFEB9C")
            vcol   = next(
                (cell.column for cell in ws[1] if cell.value == verdict_col), None
            )
            if vcol:
                for row in ws.iter_rows(min_row=2):
                    cell = row[vcol - 1]
                    if cell.value == "PASS":
                        cell.fill = green
                    elif cell.value == "FAIL":
                        cell.fill = red
                        cell.font = Font(bold=True, color="9C0006")
                    elif cell.value == "SKIP":
                        cell.fill = yellow

        # Auto-size columns
        for col in ws.columns:
            max_len = max((len(str(c.value or "")) for c in col), default=0)
            ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 60)

    return path


def read_file(path: Path) -> pd.DataFrame:
    """Read CSV or Excel into a DataFrame."""
    if path.suffix.lower() in (".xlsx", ".xls"):
        return pd.read_excel(path)
    return pd.read_csv(path)